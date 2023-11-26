import pandas as pd
from openpyxl import load_workbook
from typing import List, Any, Tuple, Dict
import openai
from split_text import split_text_by_chars
from embedding import generate_embed_openai, language_embed
from translation import NLPTranslation
import asyncio
import os
import json
from openai_components import num_tokens_from_string, generate_chat_response

# MAIN
async def Embed_spreadsheet(file_path, NLP_api_key, openai_api_key, language):
    # Validate and extract file name and extension
    file_name, file_extension = os.path.splitext(file_path)

    allowed_extensions = ['.csv', '.xlsx', '.ods', '.tsv']
    if file_extension not in allowed_extensions:
        # raise ValueError(f"Unsupported file extension. Please provide one of these types: {allowed_extensions}")
        return None

    embed_translate = await language_embed(language)
    
    if embed_translate:
        embed_language = 'en'
    else:
        embed_language = language

    tables = extract_tables(file_path)
    openai.api_key = openai_api_key

    content = {}
    idx = 0
    for table in tables:
        df = table["dataframe"]
        sheet = table["sheet"]
        result = await extract_table_content(sheet, df, NLP_api_key, openai_api_key, language, embed_translate)
        content[idx] = result  # Assuming sheet as ID
        idx += 1

    description = ""
    result_dict = await create_result(file_path, file_extension, description, language, embed_language, content)

    # Save the result as a JSON file
    with open(file_name + '.json', 'w') as fp:
        json.dump(result_dict, fp)
        
    return(result_dict)
        

#######################################################################################################################################
################################################### For dictionary Structure ##########################################################
#######################################################################################################################################

async def create_table_content(page: int, table: List[List[str]], text: str, english_text: str, embed: Any,) -> Dict[str, Any]:
    return {
        'class': "table",
        
        # Classification
        'page': page, # Which sheet
        
        # Raw table data as list of list
        'table': table,
        
        # For semantic search
        'text': text, # What is the table about?
        'english text': english_text,
        'embed': embed
    }

async def create_result(file_name: str, file_extension: str, description: str, language: str, embed_language: str, content) -> Dict[str, Any]:
    return {
        'class': "spreadsheet",

        # file name also include its file_extension
        'file name': file_name,
        'file extension': file_extension,
        
        # Language for encoding and decoding the embed text, and ai language is what language can the text pass through (default nothing as we didnt check here)
        'language': {'prefix:': language, 'ai language': "",'embed language': embed_language},
        
        # text, table, image
        'content': content
    }

#######################################################################################################################################
#################################################### For table extraction #############################################################
#######################################################################################################################################

def extract_tables(file_path: str) -> List[dict]:
    file_type = file_path.split('.')[-1]
    tables = []

    if file_type in ['csv', 'tsv']:
        delimiter = '\t' if file_type == 'tsv' else ','
        df = pd.read_csv(file_path, delimiter=delimiter)
        tables.append({
            'sheet': '0',
            'table_id': '0',
            'dataframe': df
        })
    elif file_type == 'xlsx':
        wb = load_workbook(filename=file_path, read_only=True)
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            table_id = 0
            visited = set()
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value is not None and (cell.row, cell.column) not in visited:
                        table_cells = explore_table(ws, (cell.row, cell.column), visited)
                        if table_cells:
                            table_df = create_dataframe_from_cells(ws, table_cells)
                            tables.append({
                                'sheet': sheet,
                                'table_id': str(table_id),
                                'dataframe': table_df
                            })
                            table_id += 1
    return tables

def explore_table(ws: Any, start: Tuple[int, int], visited: set) -> List[Tuple[int, int]]:
    cells_to_visit = [start]
    table_cells = []

    while cells_to_visit:
        current = cells_to_visit.pop()
        row, column = current
        table_cells.append(current)
        visited.add(current)

        # Explore neighbours (left, right, up, down)
        for dr, dc in [(-1, 0), (1, 0), (0, -1), (0, 1)]:
            neighbour = (row + dr, column + dc)
            if (neighbour[0] > 0 and neighbour[1] > 0 and  # Ensure indices are within valid range
                ws.cell(*neighbour).value is not None and 
                neighbour not in visited and
                neighbour not in cells_to_visit):
                cells_to_visit.append(neighbour)

    return table_cells

def create_dataframe_from_cells(ws: Any, cells: List[Tuple[int, int]]) -> pd.DataFrame:
    # Sort cells by their position in the spreadsheet
    cells.sort()
    data = {}
    column_names = []
    for cell in cells:
        row, column = cell
        value = ws.cell(row, column).value
        column_data = data.setdefault(column, {})
        column_data[row] = value
        # Capture column names
        if row == cells[0][0]:
            column_names.append(value)

    df = pd.DataFrame(data)
    df.sort_index(axis=1, inplace=True)
    # Set column names
    df.columns = column_names
    # Remove the row used for column names
    df.drop(cells[0][0], inplace=True)

    return df

#######################################################################################################################################
########################################################### For table content #########################################################
#######################################################################################################################################

async def extract_table_content(sheet, table, NLP_api_key, openai_api_key, language, embed_translate):
    # Validate if table meets the required conditions
    if not await is_dataframe_valid(table):
        return None
    
    # Convert the table (DataFrame) to a list of lists, including column headers.
    # This list format will be used for the final output.
    original_table = [table.columns.tolist()] + table.values.tolist()

    # Translate the table content if the 'embed_translate' flag is True
    translated_table = await translate_table(table, language, NLP_api_key) if embed_translate else table

    # Split the translated table into chunks if it exceeds a certain token limit
    chunks = await split_table_into_chunks(translated_table, token_limit=4096, column_multiplier=30)

    # Check if the full table is being used. If there is only one chunk, it implies that the full table is being used.
    full_table = len(chunks) == 1

    # Interpret each chunk of data and generate a summary and row descriptions
    output_split = await interpret_table_chunks(chunks, full_table)

    # Initialize text and english_text lists to store the processed output
    text, english_text, embeddings = [], [], []

    # Process each element in output_split
    for element in output_split:
        # Generate embeddings for the English text
        embed = await generate_embed_openai(element, openai_api_key, language) if element else {'embed': []}
        
        embeddings.append(embed["embed"])
        
        # If translation is needed, translate the text back to the original language and add to the lists
        if embed_translate:
            translated_text = (await NLPTranslation(element, NLP_api_key, "en", language))["translated text"]
            text.append(translated_text)
            english_text.append(element)
        else:
            text.append(element)
            english_text.append("")

    # Call create_table_content function to structure the final output
    table_content = await create_table_content(sheet, original_table, text, english_text, embeddings)
    return table_content

# Function to translate a table using an NLP API, if 'embed_translate' flag is True
async def translate_table(table, language, NLP_api_key):
    result = table.copy()
    
    # Generate translation tasks for non-null, non-empty and non-numeric cells
    tasks = [(row, col, await NLPTranslation(cell, NLP_api_key, language, "en")) 
             for row in table.iterrows() for col in row[1].index 
             if pd.notnull(cell := row[1][col]) and cell != "" and not isinstance(cell, (int, float))]

    # Execute translation tasks and update the result table with translated content
    for row, col, task in tasks:
        result.at[row, col] = await task

    return result

# Function to validate if a DataFrame is valid based on certain conditions
async def is_dataframe_valid(df, dot_threshold = 0.2, nan_threshold = 0.4) -> bool:
    # Checks whether the input DataFrame is valid based on two conditions:
    # 1. The percentage of periods ('.') in the string representation of the DataFrame should be below `dot_threshold`.
    # 2. The percentage of NaN values in the DataFrame should be below `nan_threshold`.

    # Concatenate all the data in the DataFrame into one string and remove spaces
    data_string_no_spaces = df.to_string().replace(" ", "")
    
    # Get the total number of cells in the DataFrame
    total_cells = df.size

    # If DataFrame is empty, return False
    if total_cells == 0 or len(data_string_no_spaces) == 0:
        return False

    # Calculate the percentage of periods and NaN values in the DataFrame
    period_percentage = data_string_no_spaces.count('.') / len(data_string_no_spaces)
    nan_percentage = df.isnull().sum().sum() / total_cells

    # Return True if both percentages are below their respective thresholds, otherwise False
    return period_percentage < dot_threshold and nan_percentage < nan_threshold

# Function to split the table into chunks if it exceeds a certain token limit
async def split_table_into_chunks(table, token_limit, column_multiplier):
    # This function works by calculating the number of tokens in each row of the table.
    
    # Create a string representing all the column names
    columns_str = "  ".join(str(column) for column in table.columns)

    # Initialize list to store token count of each row
    row_token_counts = [await num_tokens_from_string(columns_str, "cl100k_base")]

    # Calculate token count for each row and add to the list
    for _, row in table.iterrows():
        tokens = await num_tokens_from_string(str(row.values)[1:-1], "cl100k_base")
        row_token_counts.append(tokens)

    # Check if token count exceeds limit
    total_tokens = sum(row_token_counts) * (1 + column_multiplier / (table.shape[0] + table.shape[1]))
    high_token = total_tokens > token_limit

    # If it does, split the table into chunks
    if high_token:
        chunks = await create_table_chunks(table, row_token_counts, token_limit, column_multiplier)
    else:
        chunks = [table]
    return chunks

# Helper function to create chunks of table based on token limit
async def create_table_chunks(table, row_token_counts, token_limit, column_multiplier):
    # If the total number of tokens exceeds the token limit, it splits the table into chunks.

    chunks, chunk, current_token_count = [], [], 0
    for idx, row in enumerate(table.itertuples(index=False)):
        row_token = row_token_counts[idx + 1] * (1 + column_multiplier / table.shape[1])
        if current_token_count + row_token > token_limit and chunk:
            chunks.append(pd.DataFrame(chunk, columns=table.columns))
            chunk, current_token_count = [], 0
        chunk.append(dict(zip(table.columns, row)))
        current_token_count += row_token
    if chunk:
        chunks.append(pd.DataFrame(chunk, columns=table.columns))
    return chunks

# Function to interpret each chunk of the table, generate summary and row descriptions
async def interpret_table_chunks(chunks, full_table):
    # Ensure chunks is a list of pandas DataFrames
    if not isinstance(chunks, list) or not all(isinstance(chunk, pd.DataFrame) for chunk in chunks):
        raise TypeError("chunks must be a list of pandas DataFrame objects.")

    # Ensure full_table is a boolean
    if not isinstance(full_table, bool):
        raise TypeError("full_table must be a boolean.")

    # Initialize output list
    more_output = []

    # Interpret each chunk
    for chunk_df in chunks:
        chunk_string = await generate_chunk_string(chunk_df)
        
        messages = [
            {"role": "system", "content": "You are a helpful assistant."},
            {"role": "user", "content": f"""Use the table provided below to analyze the data and gain valuable insights. Focus on the numbers and identify trends that reveal important information. Highlight the most significant rows to draw attention to key findings. Create a narrative that is engaging and optimized for semantic search, ensuring that users can easily find the information they need.

Instead of using lists or bullet points, craft well-connected paragraphs that flow together seamlessly. Paint a clear and vivid picture of the data, captivating readers and encouraging them to explore further. Consider the possible questions users might ask and incorporate them naturally into your text to enhance its visibility in semantic search results.

Table: {chunk_string}"""},
        ]

        try:
            output = await generate_chat_response(messages)
            output = output["content"]
            output_split = await split_text_by_chars(output, 1000)
            
            if full_table:
                return output_split
            more_output.extend(output_split)
        except Exception as e:
            raise Exception(f"An error occurred: {e}")
    return more_output

# Helper function to generate string representation of a table chunk
async def generate_chunk_string(chunk_df):
    chunk_str = f"Column's name': {', '.join(str(chunk_df.columns))}\n"
    for i, row in chunk_df.iterrows():
        row_str = [str(element) for element in row]
        chunk_str += f"Row {i+1}: {', '.join(row_str)}\n"
    return chunk_str

